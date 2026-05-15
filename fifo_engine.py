import shutil
from collections import deque, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def run_fifo(SRC, DST):
    shutil.copy(SRC, DST)
    wb_r = load_workbook(SRC, data_only=True)
    wb   = load_workbook(DST)

    ws_inv_r = wb_r["Supplier Invoices"]
    ws_bol_r = wb_r["Purchase to BOL-RTB"]
    ws_lt_r  = wb_r["Load Tracking"]
    ws_inv   = wb["Supplier Invoices"]
    ws_bol   = wb["Purchase to BOL-RTB"]
    ws_lt    = wb["Load Tracking"]

    def weighted_avg(pairs):
        total = sum(g for g, _ in pairs)
        return sum(g * r for g, r in pairs) / total if total else 0.0

    def join_unique(items):
        seen, out = set(), []
        for x in items:
            s = str(x).strip()
            if s and s not in seen:
                seen.add(s); out.append(s)
        return " | ".join(out)

    # ══════════════════════════════════════════════════════════════════════════════
    # Build supplier invoice FIFO queues
    # Supplier Invoices (0-indexed): A(0)=Batch, C(2)=Supplier, D(3)=Invoice#,
    #   G(6)=PaidForGallons, L(11)=Rate(usd/gal), N(13)=SupplyCostDashFuel(MXN/L)
    # ══════════════════════════════════════════════════════════════════════════════
    supplier_queues = defaultdict(list)   # supplier_upper → [entry, ...]
    inv_entries     = {}                  # inv_num → entry

    for i, row in enumerate(ws_inv_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if row[0] is None: break
        supplier        = str(row[2]).strip().upper() if row[2] else ""
        inv_num         = str(row[3]).strip()         if row[3] else ""
        rate_usd_gal    = float(row[11])              if row[11] else 0.0
        rate_adder      = float(row[12])              if row[12] else 0.0
        supply_cost_mxn = float(row[13])              if row[13] else 0.0
        batch           = row[0]
        try:
            if row[6] is not None:
                paid_gals = float(row[6])
                paid_gals_real = True
            elif row[22] is not None and row[23] is not None:
                # W+X = already drawn + remainder = total paid gallons
                paid_gals = float(row[22]) + float(row[23])
                paid_gals_real = True
            elif row[4] is not None and float(row[4]) > 0:
                # Has wired amount - use Gallons as fallback
                paid_gals = float(row[5]) if row[5] is not None else 0.0
                paid_gals_real = True
            else:
                # No wired amount (e.g. Rhodes) - skip
                paid_gals = 0.0
                paid_gals_real = False
        except (TypeError, ValueError):
            paid_gals = 0.0
            paid_gals_real = False
        if paid_gals <= 0:
            continue
        entry = {
            "inv_num":         inv_num,
            "batch":           str(batch),
            "rate_usd_gal":    rate_usd_gal,
            "rate_adder":      rate_adder,
            "supply_cost_mxn": supply_cost_mxn,
            "avail":           paid_gals,
            "orig":            paid_gals,
            "drawn":           0.0,
            "excel_row":       i,
        }
        supplier_queues[supplier].append(entry)
        inv_entries[inv_num] = entry

    # ══════════════════════════════════════════════════════════════════════════════
    # Read Purchase to BOL-RTB into a lookup by BOL number
    # C(2)=Supplier, E(4)=BOL, I(8)=Gallons  (0-indexed)
    # ══════════════════════════════════════════════════════════════════════════════
    bol_info = {}   # bol_str → {supplier, gals, excel_row}
    for i, row in enumerate(ws_bol_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[2]: break
        bol_str  = str(row[5]).strip() if row[5] else ""   # F(6) BOL
        supplier = str(row[2]).strip().upper() if row[2] else ""  # C(3) Supplier
        try:
            gals = float(row[6]) if row[6] else 0.0   # G(7) Gallons
        except (TypeError, ValueError):
            gals = 0.0
        if bol_str:
            bol_info[bol_str] = {"supplier": supplier, "gals": gals, "excel_row": i}

    # ══════════════════════════════════════════════════════════════════════════════
    # STAGE 1 — Allocate supplier invoices to BOLs
    # Order: Load Tracking RTB/RTC rows (when available) → else Purchase to BOL-RTB order
    # This supports both workflows:
    #   Upload 1: User fills Supplier Invoices + Purchase to BOL-RTB only (no LT yet)
    #             → allocate in Purchase to BOL-RTB row order
    #   Upload 2: User adds RTBs to Load Tracking
    #             → allocate in Load Tracking order (master FIFO order)
    # ══════════════════════════════════════════════════════════════════════════════
    bol_alloc = {}   # bol_str → {inv_str, batch_str, cost_usd, cost_mxn}

    # Determine allocation order
    lt_rtb_bols = []  # BOLs found in Load Tracking RTB/RTC rows, in sheet order
    for i, row in enumerate(ws_lt_r.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not row[0]: continue
        if row[10] not in ("RTB", "RTC"): continue
        bol_str = str(row[30]).strip() if row[30] else ""
        if bol_str and bol_str in bol_info:
            lt_rtb_bols.append(bol_str)

    # If no RTBs in Load Tracking yet, use Purchase to BOL-RTB row order
    if lt_rtb_bols:
        alloc_order = lt_rtb_bols
    else:
        alloc_order = [str(row[5]).strip() for _, row in
                       enumerate(ws_bol_r.iter_rows(values_only=True), start=1)
                       if _ > 7 and row[2] and row[5]]

    for bol_str in alloc_order:
        info = bol_info.get(bol_str)
        if not info:
            continue

        supplier = info["supplier"]
        bol_gals = info["gals"]
        if bol_gals <= 0:
            continue

        queue = supplier_queues.get(supplier, [])
        if not queue:
            continue
        remaining = bol_gals
        alloc_usd, alloc_adder, alloc_mxn, inv_labels, batch_labels = [], [], [], [], []

        for inv in queue:
            if remaining <= 1e-6: break
            if inv["avail"] <= 1e-6: continue
            draw = min(remaining, inv["avail"])
            alloc_usd.append((draw, inv["rate_usd_gal"]))
            alloc_adder.append((draw, inv["rate_adder"]))
            alloc_mxn.append((draw, inv["supply_cost_mxn"]))
            if inv["inv_num"] not in inv_labels:
                inv_labels.append(inv["inv_num"])
            if inv["batch"] not in batch_labels:
                batch_labels.append(inv["batch"])
            inv["avail"] -= draw
            inv["drawn"] += draw
            remaining    -= draw

        if remaining > 1e-6:
            supplier_name = supplier.title()
            raise ValueError(
                f"Not enough allocation for BOL {bol_str} ({supplier_name}): "
                f"need {bol_gals:.2f} gals, short by {remaining:.2f} gals. "
                f"Please add more paid-for gallons in Supplier Invoices."
            )

        bol_alloc[bol_str] = {
            "inv_str":   join_unique(inv_labels),
            "batch_str": join_unique(batch_labels),
            "cost_usd":    weighted_avg(alloc_usd),
            "cost_adder":  weighted_avg(alloc_adder),
            "cost_mxn":    weighted_avg(alloc_mxn),
        }

    # ══════════════════════════════════════════════════════════════════════════════
    # Write Stage 1 results back to Purchase to BOL-RTB (in its own row order)
    # D(4)=SupplierInvoice, J(10)=Cost/GalUSD, R(18)=Total/LiterMXN
    # ══════════════════════════════════════════════════════════════════════════════
    # Purchase to BOL-RTB column layout (new):
    #   D(4)=Supplier Invoice, E(5)=Batch, F(6)=BOL, G(7)=Gallons,
    #   I(9)=Cost/Gal USD [WRITE], K(11)=Cost/Gal+Adder [WRITE]
    #   J(10)=Adder [formula=K-I], N(14)=Freight/Gal [formula=M/G],
    #   O(15)=Total Cost/Gal [formula=N+K], Q(17)=Invoice Amount [formula=O*G],
    #   T(20)=Balance [formula], U-W=flags [formula] — all left as formulas
    for i, row in enumerate(ws_bol_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[2]: break
        bol_str = str(row[5]).strip() if row[5] else ""  # F(6) BOL, index 5
        alloc   = bol_alloc.get(bol_str, {})
        if not alloc: continue
        ws_bol.cell(row=i, column=4).value  = alloc["inv_str"]      # D Supplier Invoice
        ws_bol.cell(row=i, column=5).value  = alloc["batch_str"]   # E Batch
        ws_bol.cell(row=i, column=10).value = round(alloc["cost_usd"],    6)  # J Cost/Gal USD
        ws_bol.cell(row=i, column=12).value = round(alloc["cost_adder"],  6)  # L Cost/Gal+Adder
        ws_bol.cell(row=i, column=13).value = round(alloc["cost_mxn"],    6)  # M Supply Cost DashFuel (MXN/L)

    # Write Net RTB Gallons, Remainder, and Liter formulas to Supplier Invoices
    # W(23)=NetRTBGallons, X(24)=RemainderGallons, U(21)=formula, V(22)=formula
    for inv in inv_entries.values():
        rem = inv["orig"] - inv["drawn"]
        r   = inv["excel_row"]
        ws_inv.cell(row=r, column=23).value = round(inv["drawn"], 6)
        ws_inv.cell(row=r, column=24).value = round(rem, 6)
        ws_inv.cell(row=r, column=21).value = f"=W{r}*3.7854"
        ws_inv.cell(row=r, column=22).value = f"=X{r}*3.7854"

    # ══════════════════════════════════════════════════════════════════════════════
    # STAGE 2 — RTB → BTC FIFO on Load Tracking (sheet order = master order)
    # Load Tracking (0-indexed): K(10)=Groups, L(11)=LocationName, O(14)=LocCity,
    #   P(15)=Product, W(22)=NetLiters, Z(25)=TerminalName, AE(30)=BOL,
    #   AR(43)=SupplyCost, AV(47)=TotalCost/L
    # Write-back (1-indexed): AR(44), BE(57)=Batch, BF(58)=SupplierInv, BG(59)=BOLSource
    # ══════════════════════════════════════════════════════════════════════════════
    ws_lt.cell(row=1, column=59).value = "BOL Source"
    ws_lt.cell(row=1, column=60).value = "Batch Source"

    inventory      = {}   # (product, bulk_plant) → deque of slots
    fifo_log       = []
    bol_remaining  = {}   # bol_str → remaining liters (updated as BTCs consume)

    for i, row in enumerate(ws_lt_r.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not row[0]: continue

        grp           = row[10]
        product       = str(row[15]).strip() if row[15] else "Unknown"
        location_name = str(row[11]).strip() if row[11] else "Unknown"
        terminal_name = str(row[25]).strip().lstrip("* ") if row[25] else "Unknown"
        customer      = str(row[8]).strip()  if row[8]  else ""
        location_city = str(row[14]).strip() if row[14] else ""
        ld_num        = row[0]
        bol           = str(row[30]).strip() if row[30] else ""
        pickup        = row[3]

        try:
            net_liters = float(row[22]) if row[22] else 0.0
        except (TypeError, ValueError):
            net_liters = 0.0

        bulk_plant = location_name if grp == "RTB" else terminal_name
        key = (product, bulk_plant)

        if grp == "RTB":
            supply_cost  = float(row[43]) if row[43] else 0.0
            av_available = row[47] is not None
            total_cost_l = float(row[47]) if row[47] else (float(row[43]) if row[43] else 0.0)
            alloc        = bol_alloc.get(bol, {})
            batch_str    = alloc.get("batch_str", "")
            inv_str      = alloc.get("inv_str", "")

            if key not in inventory:
                inventory[key] = deque()
            supplier_upper = str(row[28]).strip().upper() if row[28] else ""
            inventory[key].append({
                "liters": net_liters, "cost": total_cost_l, "av_available": av_available,
                "bol": bol, "batch": batch_str, "inv": inv_str,
                "supplier_upper": supplier_upper,
            })
            bol_remaining[bol] = net_liters

            ws_lt.cell(row=i, column=44).value = supply_cost
            ws_lt.cell(row=i, column=57).value = batch_str   # BE Batch
            ws_lt.cell(row=i, column=58).value = inv_str     # BF Supplier Invoice
            ws_lt.cell(row=i, column=59).value = ""          # BG BOL Source
            ws_lt.cell(row=i, column=60).value = ""          # BH Batch Source

            queue_rem = sum(s["liters"] for s in inventory[key])
            fifo_log.append({
                "type": "RTB", "ld": ld_num, "pickup": pickup,
                "product": product, "bulk_plant": bulk_plant,
                "customer": customer, "bol": bol,
                "batch": batch_str, "liters": net_liters, "cost_per_l": total_cost_l,
                "total_cost": net_liters * total_cost_l,
                "source_bols": "-", "queue_rem": queue_rem,
                "remaining_l": net_liters,  # will be updated after BTCs consume
            })

        elif grp == "RTC":
            alloc     = bol_alloc.get(bol, {})
            ws_lt.cell(row=i, column=57).value = alloc.get("batch_str", "")  # BE Batch
            ws_lt.cell(row=i, column=58).value = alloc.get("inv_str", "")    # BF Supplier Invoice
            ws_lt.cell(row=i, column=59).value = ""                           # BG BOL Source
            ws_lt.cell(row=i, column=60).value = ""                           # BH Batch Source

        elif grp == "BTC":
            remaining, allocations = net_liters, []
            source_bols, source_batches = [], []

            q = inventory.get(key, deque())
            while remaining > 1e-6 and q:
                slot = q[0]
                draw = min(remaining, slot["liters"])
                allocations.append((draw, slot["cost"]))
                if slot["bol"] not in source_bols:
                    source_bols.append(slot["bol"])
                for b in slot["batch"].split(" | "):
                    b = b.strip()
                    if b and b not in source_batches:
                        source_batches.append(b)
                slot["liters"] -= draw
                bol_remaining[slot["bol"]] = slot["liters"]   # update remaining
                remaining      -= draw
                if slot["liters"] <= 1e-6:
                    q.popleft()

            if remaining > 1e-6:
                allocations.append((remaining, 0.0))
                source_bols.append("No RTB")

            cost_per_l = weighted_avg(allocations)
            bols_str   = join_unique(source_bols)
            batch_str  = join_unique(source_batches)
            queue_rem  = sum(s["liters"] for s in inventory.get(key, deque()))

            # If source RTBs had no AV (Total Cost/L), fall back to existing BTC AR
            existing_ar = float(row[43]) if row[43] else 0.0
            any_av = any(s.get("av_available", False) for s in (list(inventory.get(key, [])) or []))
            used_fallback = not any_av and existing_ar > 0
            final_cost  = existing_ar if used_fallback else round(cost_per_l, 6)
            ws_lt.cell(row=i, column=44).value = final_cost
            ws_lt.cell(row=i, column=57).value = ""           # BE Batch (blank for BTC)
            ws_lt.cell(row=i, column=58).value = ""           # BF Supplier Invoice (blank for BTC)
            ws_lt.cell(row=i, column=59).value = bols_str     # BG BOL Source
            ws_lt.cell(row=i, column=60).value = batch_str    # BH Batch Source

            fifo_log.append({
                "type": "BTC", "ld": ld_num, "pickup": pickup,
                "product": product, "bulk_plant": bulk_plant,
                "customer": customer, "bol": bol,
                "batch": batch_str, "liters": net_liters, "cost_per_l": cost_per_l,
                "total_cost": net_liters * cost_per_l,
                "source_bols": bols_str, "queue_rem": queue_rem,
                "remaining_l": None,  # BTCs have no remaining — they consume
            })

    # Backfill remaining_l for RTB entries from bol_remaining
    for entry in fifo_log:
        if entry["type"] == "RTB":
            entry["remaining_l"] = bol_remaining.get(entry["bol"], entry["liters"])

    # ══════════════════════════════════════════════════════════════════════════════
    # FIFO sheet
    # ══════════════════════════════════════════════════════════════════════════════
    FILL_RTB    = PatternFill("solid", fgColor="C6EFCE")
    FILL_BTC    = PatternFill("solid", fgColor="FFDDC1")
    FILL_HDR    = PatternFill("solid", fgColor="2F5496")
    FILL_SUBHDR = PatternFill("solid", fgColor="BDD7EE")
    thin        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, val):
        cell.value = val; cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    def sub(cell, val):
        cell.value = val; cell.fill = FILL_SUBHDR
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    def dat(cell, val, num_fmt=None, fill=None):
        cell.value = val; cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        if num_fmt: cell.number_format = num_fmt
        if fill:    cell.fill = fill

    if "FIFO" in wb.sheetnames:
        del wb["FIFO"]
    ws_f = wb.create_sheet("FIFO")

    COLS = [
        ("Load #",             11), ("Date",              12), ("Type",           7),
        ("Product",            12), ("Bulk Plant",        12), ("Customer",       22),
        ("Batch",               8), ("BOL",               12), ("Liters",         14),
        ("Cost / L (MXN)",     15), ("Total Cost (MXN)",  17), ("Source BOLs",    36),
        ("Queue Bal. (L)",     17), ("Remaining L (BOL)", 18),
    ]

    for ci, (name, width) in enumerate(COLS, start=1):
        hdr(ws_f.cell(row=1, column=ci), name)
        ws_f.column_dimensions[get_column_letter(ci)].width = width
    ws_f.row_dimensions[1].height = 32

    running = {}
    for ri, entry in enumerate(fifo_log, start=2):
        rkey = (entry["product"], entry["bulk_plant"])
        fill = FILL_RTB if entry["type"] == "RTB" else FILL_BTC
        if rkey not in running: running[rkey] = 0.0
        running[rkey] += entry["liters"] if entry["type"] == "RTB" else -entry["liters"]
        for ci, (val, fmt) in enumerate([
            (entry["ld"],            None),
            (entry["pickup"],        "DD/MM/YYYY"),
            (entry["type"],          None),
            (entry["product"],       None),
            (entry["bulk_plant"],    None),
            (entry["customer"],      None),
            (entry["batch"],         None),
            (entry["bol"],           None),
            (entry["liters"],        "#,##0.00"),
            (entry["cost_per_l"],    "#,##0.0000"),
            (entry["total_cost"],    "$#,##0.00"),
            (entry["source_bols"],   None),
            (running[rkey],          "#,##0.00"),
            (entry["remaining_l"],       "#,##0.00"),
        ], start=1):
            dat(ws_f.cell(row=ri, column=ci), val, num_fmt=fmt, fill=fill)
        ws_f.row_dimensions[ri].height = 16

    ws_f.freeze_panes = "A2"

    sr = len(fifo_log) + 3
    ws_f.cell(row=sr, column=1).value = "INVENTORY REMAINING IN QUEUE"
    ws_f.cell(row=sr, column=1).font  = Font(bold=True, size=10)
    sr += 1
    for ci, label in enumerate(["Product", "Bulk Plant", "Liters in Queue", "Next Cost/L (MXN)", "Avg Cost in Inventory (MXN/L)"], start=1):
        sub(ws_f.cell(row=sr, column=ci), label)
    for (prod, bp), q in inventory.items():
        if not q: continue
        sr += 1
        dat(ws_f.cell(row=sr, column=1), prod,                        fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=2), bp,                          fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=3), sum(s["liters"] for s in q), num_fmt="#,##0.00",   fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=4), q[0]["cost"],                num_fmt="#,##0.0000", fill=FILL_SUBHDR)

    # ══════════════════════════════════════════════════════════════════════════════
    # FIFO sheet — Average Cost in Remaining Inventory (col E of summary block)
    # ══════════════════════════════════════════════════════════════════════════════
    # inventory dict is still in scope: (product, bulk_plant) → deque of slots
    # The summary block starts at row (len(fifo_log) + 4); data rows begin 2 after that.
    # We need to find and fill col E for each inventory summary row.
    # Re-derive the summary start row the same way the sheet builder did.
    summary_data_start = len(fifo_log) + 5   # row 26=label, 27=header, 28+=data

    inv_items = [(k, q) for k, q in inventory.items() if q]
    for idx, ((prod, bp), q) in enumerate(inv_items):
        row_num = summary_data_start + idx
        total_l   = sum(s["liters"] for s in q)
        avg_cost  = sum(s["liters"] * s["cost"] for s in q) / total_l if total_l else 0.0
        ws_f.cell(row=row_num, column=5).value = round(avg_cost, 6)
        # also style it to match the other summary cells
        dat(ws_f.cell(row=row_num, column=5), round(avg_cost, 6),
            num_fmt="#,##0.0000", fill=FILL_SUBHDR)

    # ══════════════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════════
    # Overall Summary — all columns B through K
    # B=Total Invoices USD, C=Total Gallons, D=Total Wired, E=Paid for Gallons,
    # F=Gallons Pulled, G=Remaining in Allocation, H=Remaining in Inventory,
    # I=Weighted Avg Cost, J=Amount Paid Back Mexico, K=Mexico Balance
    # J and K read from ws_bol (already written with correct Cost/Gal)
    # ══════════════════════════════════════════════════════════════════════════════
    LITERS_PER_GAL = 3.7854
    from collections import defaultdict as _dd

    # Aggregate Overall Summary values
    # B (Total Invoice USD), C (Total Gallons), E (Paid for Gallons):
    #   Read directly from source Overall Summary which has cached formula values
    # D (Total Wired): read from Supplier Invoices col E (always raw)
    # F (Gallons Pulled), G (Remaining Allocation): from Supplier Invoices W/X (script-written)
    sup_inv = _dd(lambda: {"b":0.0,"c":0.0,"d":0.0,"e":0.0,"f":0.0,"g":0.0})

    def _f(v):
        try: return float(v) if v is not None else 0.0
        except: return 0.0

    # Read B, C, E from source Overall Summary (cached formula values)
    ws_os_r = wb_r["Overall Summary"]
    for i, row in enumerate(ws_os_r.iter_rows(values_only=True), start=1):
        if i < 4: continue
        label = str(row[0]).strip() if row[0] else ""
        if not label or label == "Row Labels": continue
        if "TOTAL" in label.upper(): continue
        sup_inv[label]["b"] = _f(row[1])   # B Total Invoice USD (formula cached)
        sup_inv[label]["c"] = _f(row[2])   # C Total Gallons
        sup_inv[label]["e"] = _f(row[4])   # E Paid for Gallons (formula cached)
        sup_inv[label]["h_src"] = _f(row[7])  # H Rem Inventory (cached, fallback)
        sup_inv[label]["i_src"] = _f(row[8])  # I Avg Cost (cached, fallback)

    # Read D (Wired Amount) from Supplier Invoices col E (raw)
    # Read F (Pulled) and G (Remaining) from script-written W/X
    out_rows = {i: row for i, row in enumerate(ws_inv.iter_rows(values_only=True), start=1)
                if i > 7 and row[0] is not None}
    for i, out in out_rows.items():
        sup = str(out[2]).strip()
        sup_inv[sup]["d"] += _f(out[4])    # E Wired Amount (raw)
        sup_inv[sup]["f"] += _f(out[22])   # W Net RTB Gallons (script-written)
        sup_inv[sup]["g"] += _f(out[23])   # X Remainder Gallons (script-written)

    # Aggregate J (Received Payments) and K (Mexico Balance) from BOL output
    # T(Balance) is a formula — compute it: T = (M/G + K)*G - S if P!="" else (M/G+K)*G
    # Using ws_bol (output) so I and K are already written by the script
    sup_j = _dd(float)
    sup_k = _dd(float)
    for i, row in enumerate(ws_bol.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[2]: break
        sup_raw = str(row[2]).strip()
        try:
            recv = float(row[18]) if row[18] is not None else 0.0  # S(19) Received Payments
        except (TypeError, ValueError):
            recv = 0.0
        # Compute Balance directly from raw values (avoids formula evaluation issue)
        try:
            gals    = float(row[6])  if row[6]  is not None else 0.0  # G(7) Gallons
            freight = float(row[12]) if row[12] is not None else 0.0  # M(13) Freight/Load
            cost_k  = float(row[10]) if row[10] is not None else 0.0  # K(11) Cost/Gal+Adder
            inv_num = row[17]                                           # R(18) Invoice#
            inv_amt = (freight + cost_k * gals) if gals else 0.0       # Q = O*G = (N+K)*G
            bal     = (inv_amt - recv) if inv_num else inv_amt
        except (TypeError, ValueError):
            bal = 0.0
            inv_num = None
        sup_j[sup_raw] += recv
        if bal > 0.01 and inv_num:
            sup_k[sup_raw] += bal

    # Remaining FIFO inventory by supplier
    supplier_remaining = _dd(list)
    for (prod, bp), q in inventory.items():
        for slot in q:
            sup_upper = slot.get("supplier_upper", "")
            if sup_upper:
                supplier_remaining[sup_upper].append((slot["liters"], slot["cost"]))

    def _match(label, d):
        lu = label.upper()
        total = 0.0
        for k, v in d.items():
            if k.upper() in lu or lu in k.upper():
                total += v
        return total

    def _match_dict(label, d):
        lu = label.upper()
        for k, v in d.items():
            if k.upper() in lu or lu in k.upper():
                return v
        return {"b":0.0,"c":0.0,"d":0.0,"e":0.0,"f":0.0,"g":0.0}

    def _match_slots(label, d):
        lu = label.upper()
        result = []
        for k, v in d.items():
            if k.upper() in lu or lu in k.upper():
                result.extend(v)
        return result

    ws_os = wb["Overall Summary"]
    gt = {"b":0.0,"c":0.0,"d":0.0,"e":0.0,"f":0.0,"g":0.0,
          "h_l":0.0,"h_pairs":[],"j":0.0,"k":0.0}

    for i, row in enumerate(ws_os.iter_rows(values_only=False), start=1):
        if i < 4: continue
        label = str(row[0].value).strip() if row[0].value else ""
        if not label or label == "Row Labels":
            continue

        if "TOTAL" in label.upper():
            h_gals = gt["h_l"] / LITERS_PER_GAL
            i_avg  = (sum(l*c for l,c in gt["h_pairs"]) / sum(l for l,_ in gt["h_pairs"])
                      if gt["h_pairs"] else 0.0)
            # If no real AV-based costs, use source cached Total row H and I
            slots_have_av = any(s.get("av_available", True) for s in
                                [slot for (prod,bp),q in inventory.items() for slot in q])
            if not slots_have_av:
                # Find cached total row in source Overall Summary
                for src_row in ws_os_r.iter_rows(values_only=True):
                    if src_row[0] and "TOTAL" in str(src_row[0]).upper():
                        if src_row[7]: h_gals = float(src_row[7])
                        if src_row[8]: i_avg  = float(src_row[8])
                        break
            ws_os.cell(row=i, column=2).value  = round(gt["b"], 6)
            ws_os.cell(row=i, column=3).value  = round(gt["c"], 6)
            ws_os.cell(row=i, column=4).value  = round(gt["d"], 6)
            ws_os.cell(row=i, column=5).value  = round(gt["e"], 6)
            ws_os.cell(row=i, column=6).value  = round(gt["f"], 6)
            ws_os.cell(row=i, column=7).value  = round(gt["g"], 6)
            ws_os.cell(row=i, column=8).value  = round(h_gals, 6)
            ws_os.cell(row=i, column=9).value  = round(i_avg, 6)
            ws_os.cell(row=i, column=10).value = round(gt["j"], 6)
            ws_os.cell(row=i, column=11).value = round(gt["k"], 6)
            break

        sd     = _match_dict(label, sup_inv)
        j_v    = _match(label, sup_j)
        k_v    = _match(label, sup_k)
        slots  = _match_slots(label, supplier_remaining)
        h_l    = sum(l for l,_ in slots)
        h_gals = h_l / LITERS_PER_GAL
        i_avg  = (sum(l*c for l,c in slots)/h_l if h_l else 0.0)
        # If FIFO slots used fallback costs (AV not cached), use source cached H/I
        slots_have_av = any(s.get("av_available", True) for s in
                            [slot for (prod,bp),q in inventory.items()
                             for slot in q])
        if not slots_have_av and sd.get("i_src", 0) > 0:
            i_avg  = sd["i_src"]
        if not slots_have_av and sd.get("h_src", 0) > 0:
            h_gals = sd["h_src"]

        ws_os.cell(row=i, column=2).value  = round(sd["b"], 6)
        ws_os.cell(row=i, column=3).value  = round(sd["c"], 6)
        ws_os.cell(row=i, column=4).value  = round(sd["d"], 6)
        ws_os.cell(row=i, column=5).value  = round(sd["e"], 6)
        ws_os.cell(row=i, column=6).value  = round(sd["f"], 6)
        ws_os.cell(row=i, column=7).value  = round(sd["g"], 6)
        ws_os.cell(row=i, column=8).value  = round(h_gals, 6)
        ws_os.cell(row=i, column=9).value  = round(i_avg, 6)
        ws_os.cell(row=i, column=10).value = round(j_v, 6)
        ws_os.cell(row=i, column=11).value = round(k_v, 6)

        gt["b"] += sd["b"]; gt["c"] += sd["c"]; gt["d"] += sd["d"]
        gt["e"] += sd["e"]; gt["f"] += sd["f"]; gt["g"] += sd["g"]
        gt["h_l"] += h_l;  gt["h_pairs"].extend(slots)
        gt["j"] += j_v;    gt["k"] += k_v

    wb.save(DST)
    print("Done")



if __name__ == "__main__":
    run_fifo(
        "/mnt/user-data/uploads/Investor_Summary_FIFO_7.xlsx",
        "/mnt/user-data/outputs/Investor_Summary_FIFO.xlsx"
    )
