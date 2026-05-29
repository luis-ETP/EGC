"""
Extract structured data from a processed FIFO workbook for the dashboard.
v2 - includes investment summary extraction
"""
import openpyxl
from collections import defaultdict

LITERS_PER_GAL = 3.7854

def extract(path, src_path=None):
    wb     = openpyxl.load_workbook(path, data_only=True)
    wb_src = openpyxl.load_workbook(src_path, data_only=True) if src_path else wb

    overall_summary = _extract_overall_summary(wb_src, wb_fifo=wb)
    inventory       = _extract_inventory(wb)
    fifo_rows       = _extract_fifo(wb)
    meta            = _extract_meta(wb_src, wb_fifo=wb)

    bol_tab         = _extract_bol(wb, wb_src=wb_src)
    overview_exp    = _extract_overview(wb_src)
    investment      = _extract_investment_summary(wb_src, wb_fifo=wb)
    return overall_summary, inventory, fifo_rows, meta, investment, bol_tab, overview_exp

# ── Overall Summary ────────────────────────────────────────────────────────────
def _extract_overall_summary(wb, wb_fifo=None):
    # Try reading from Overall Summary tab first (legacy)
    try:
        ws = wb["Overall Summary"]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = next((i for i, r in enumerate(rows) if r[0] == "Row Labels"), None)
        if hdr_idx is not None:
            headers = [str(v).strip() if v else f"col{j}" for j, v in enumerate(rows[hdr_idx])]
            result = []
            for row in rows[hdr_idx + 1:]:
                if not row[0]: continue
                entry = {}
                for j, h in enumerate(headers):
                    v = row[j]
                    entry[h] = float(v) if isinstance(v, (int, float)) else (str(v) if v else None)
                entry["_wired"]     = entry.get("Total Wired Amount", 0) or 0
                entry["_paid_gal"]  = entry.get("Paid for Gallons (Allocation)") or entry.get("Paid for Gallons ") or 0
                entry["_pulled"]    = entry.get("Gallons Pulled from Allocation (RTB & RTC)") or entry.get("Gallons Pulled (RTB & RTC)") or 0
                entry["_rem_alloc"] = entry.get("Remaining Gallons in Allocation") or 0
                entry["_rem_inv"]   = entry.get("Remaining Gallons in Inventory") or 0
                entry["_avg_cost"]  = entry.get("Weighted Average Cost in Inventory (MXN/L)") or entry.get("Weighted Average Cost in Inventory") or 0
                entry["_paid_back"] = entry.get("Amount Paid Back by Mexico (MXN)") or entry.get("Amount Paid Back by Mexico ") or entry.get("Amount Paid Back by Mexico") or 0
                entry["_balance"]   = entry.get("Mexico Balance (MXN)") or entry.get("Mexico Balance") or 0
                result.append(entry)
            if result:
                return result
    except KeyError:
        pass

    # Compute from raw sheets when Overall Summary tab is missing
    from collections import defaultdict
    f = lambda v: float(v) if isinstance(v, (int, float)) else 0.0

    # Supplier Invoices → wired, paid_gal per supplier
    sup_data = defaultdict(lambda: {"wired": 0.0, "paid_gal": 0.0})
    try:
        ws_si = wb["Supplier Invoices"]
        si_rows = list(ws_si.iter_rows(values_only=True))
        si_hdr = next((i for i, r in enumerate(si_rows) if r[0] == "Batch" and len(r) > 2 and r[2] == "Supplier"), None)
        if si_hdr is not None:
            sc = {str(v).strip(): j for j, v in enumerate(si_rows[si_hdr]) if v}
            for row in si_rows[si_hdr + 1:]:
                if not any(row): break
                s = str(row[sc.get("Supplier", 2)] or "").strip()
                if not s or s == "Total": continue
                sup_data[s]["wired"]    += f(row[sc.get("Wired Amount", 4)])
                sup_data[s]["paid_gal"] += f(row[sc.get("Paid for Gallons", 6)])
    except Exception:
        pass

    si_names = {s.upper(): s for s in sup_data}
    def _match(raw): return si_names.get(raw.upper(), raw)

    # Pulled gallons: sum Gallons from Purchase to BOL-RTB for every row with a BOL.
    # This is the source of truth — replaces Supplier Invoices col W (Net RTB Gallons).
    pulled_gal  = defaultdict(float)
    rem_inv_gal = defaultdict(float)
    rem_inv_mxn = defaultdict(float)
    bol_sup = {}  # BOL → raw supplier name (for FIFO remaining lookup)

    # Load Purchase to BOL-RTB once — reused for both pulled_gal and rem_alloc blocks
    bp_rows = []
    bp_hdr_idx = 6
    col_bp_bol = col_bp_sup = col_bp_gal = col_bp_inv = None
    try:
        ws_bol_pulled = wb["Purchase to BOL-RTB"]
        bp_rows = list(ws_bol_pulled.iter_rows(values_only=True))
        for i, row in enumerate(bp_rows):
            if row[0] and 'DashFuel' in str(row[0]):
                bp_hdr_idx = i
                break
        else:
            for i, row in enumerate(bp_rows):
                if any('DashFuel' in str(v) for v in row if v):
                    bp_hdr_idx = i
                    break
        bp_headers = bp_rows[bp_hdr_idx]
        def _bpcol(name, default=None):
            for j, h in enumerate(bp_headers):
                if h and name.lower() in str(h).lower():
                    return j
            return default
        col_bp_bol = _bpcol('BOL', 5)
        col_bp_sup = _bpcol('Supplier', 2)
        col_bp_gal = _bpcol('Gallons', 6)
        col_bp_inv = _bpcol('Supplier Invoice', 3)
    except Exception:
        pass

    try:
        for row in bp_rows[bp_hdr_idx + 1:]:
            if not any(row): break
            bol_val = row[col_bp_bol] if col_bp_bol is not None else None
            if not bol_val: continue  # only rows with a BOL = pulled
            raw_s = str(row[col_bp_sup] or "").strip() if col_bp_sup is not None else ""
            s     = _match(raw_s)
            if not s: continue
            bol_sup[str(bol_val)] = raw_s
            pulled_gal[s] += f(row[col_bp_gal]) if col_bp_gal is not None else 0.0
    except Exception:
        pass

    # Remaining in allocation = paid (SI) - pulled (BOL-RTB) per supplier
    # Wtd Avg Rate = Rate (usd/gal) from SI, weighted by remaining gallons per invoice
    # For split-invoice BOLs, assign gallons to the LAST invoice in the split.
    # NOTE: "Paid for Gallons" is a formula in the FIFO output — always read SI from wb (source).
    rem_alloc_gal = defaultdict(float)
    wtd_rate_num  = defaultdict(float)
    wtd_rate_den  = defaultdict(float)
    try:
        ws_si3 = wb["Supplier Invoices"]  # wb = source workbook in this function
        si_rows3 = list(ws_si3.iter_rows(values_only=True))
        si_hdr3 = next((i for i, r in enumerate(si_rows3) if r[0] == "Batch" and len(r) > 2 and r[2] == "Supplier"), None)
        if si_hdr3 is not None:
            sc3 = {str(v).strip(): j for j, v in enumerate(si_rows3[si_hdr3]) if v}
            col_sup3  = sc3.get("Supplier", 2)
            col_inv3  = sc3.get("Invoice #", 3)
            col_paid3 = sc3.get("Paid for Gallons", 6)
            col_rate3 = sc3.get("Rate (usd/gal)", 11)

            # paid and rate per (sup_upper, inv_upper) — sum paid if same invoice appears multiple times
            inv_paid3 = defaultdict(float)
            inv_rate3 = {}
            for row in si_rows3[si_hdr3 + 1:]:
                if not any(row): break
                s   = str(row[col_sup3] or "").strip()
                inv = str(row[col_inv3] or "").strip()
                if not s or s == "Total": continue
                key = (s.upper(), inv.upper())
                inv_paid3[key] += f(row[col_paid3])
                inv_rate3[key]  = f(row[col_rate3])  # rate same across dupes

            # pulled per (sup_upper, inv_upper) from BOL-RTB
            # split-invoice BOLs: assign all gallons to the LAST invoice in the split
            pulled_inv3 = defaultdict(float)
            for row in bp_rows[bp_hdr_idx + 1:]:
                if not any(row): break
                bol_val = row[col_bp_bol] if col_bp_bol is not None else None
                if not bol_val: continue
                s   = str(row[col_bp_sup] or "").strip().upper() if col_bp_sup is not None else ""
                inv = str(row[col_bp_inv] or "").strip() if col_bp_inv is not None else ""
                gal = f(row[col_bp_gal]) if col_bp_gal is not None else 0.0
                parts = [p.strip().upper() for p in inv.split("|")]
                pulled_inv3[(s, parts[-1])] += gal

            # remaining per invoice → rem_alloc and wtd_rate per supplier
            for (s_up, inv_up), paid in inv_paid3.items():
                pulled = pulled_inv3.get((s_up, inv_up), 0.0)
                rem  = max(0.0, paid - pulled)
                rate = inv_rate3.get((s_up, inv_up), 0.0)
                s_canonical = _match(s_up)
                rem_alloc_gal[s_canonical] += rem
                if rem > 0 and rate > 0:
                    wtd_rate_num[s_canonical] += rate * rem
                    wtd_rate_den[s_canonical] += rem
    except Exception:
        pass

    _wb_f = wb_fifo if wb_fifo is not None else wb
    try:
        ws_fifo = _wb_f["FIFO"]
        fh = {str(v).strip(): j for j, v in enumerate(next(ws_fifo.iter_rows(values_only=True))) if v}
        for row in ws_fifo.iter_rows(min_row=2, values_only=True):
            if not row[0]: break
            if row[fh["Type"]] != "RTB": continue
            s = _match(bol_sup.get(str(row[fh["BOL"]]), ""))
            remaining = f(row[fh["Remaining L (BOL)"]])
            cost_l    = f(row[fh["Cost / L (MXN)"]])
            rem_inv_gal[s] += remaining / 3.7854
            rem_inv_mxn[s] += remaining * cost_l
    except Exception:
        pass

    # BOL sheet → Mexico payments
    # Open balance = only rows that HAVE an invoice number (actual open invoices)
    total_received = total_balance = 0.0
    try:
        ws_bol_rtb = wb["Purchase to BOL-RTB"]
        bol_rows = list(ws_bol_rtb.iter_rows(values_only=True))
        bh = {str(v).strip(): j for j, v in enumerate(bol_rows[6]) if v}
        for row in bol_rows[7:]:
            if not row[0]: continue
            inv_num = row[bh.get("Invoice #", 17)]
            total_received += f(row[bh.get("Received Payments", 20)])
            if inv_num:  # only invoiced rows count as open balance
                total_balance += f(row[bh.get("Balance", 21)])
    except Exception:
        pass

    # Per-supplier paid_back and balance from BOL sheet
    sup_paid_back = defaultdict(float)
    sup_balance   = defaultdict(float)
    try:
        ws_bol2 = wb["Purchase to BOL-RTB"]
        bol_rows2 = list(ws_bol2.iter_rows(values_only=True))
        bh2 = {str(v).strip(): j for j, v in enumerate(bol_rows2[6]) if v}
        for row in bol_rows2[7:]:
            if not row[0]: continue
            raw_s = str(row[bh2.get("Supplier", 2)] or "").strip()
            s = _match(raw_s)
            recv_usd = f(row[bh2.get("Received Payments", 20)])
            inv_num  = row[bh2.get("Invoice #", 17)]
            bal_usd  = f(row[bh2.get("Balance", 21)]) if inv_num else 0.0
            # Received and Balance are already in USD
            if recv_usd > 0:
                sup_paid_back[s] += recv_usd
            if bal_usd > 0 and inv_num:
                sup_balance[s] += bal_usd
    except Exception:
        pass

    result = []
    for s in sup_data:
        d = sup_data[s]
        pulled   = pulled_gal.get(s, 0.0)
        rem_inv  = rem_inv_gal.get(s, 0.0)
        rem_alloc = rem_alloc_gal.get(s, 0.0)
        inv_mxn  = rem_inv_mxn.get(s, 0.0)
        avg_cost = (inv_mxn / (rem_inv * 3.7854)) if rem_inv > 0 else 0.0
        wtd_rate = (wtd_rate_num.get(s, 0.0) / wtd_rate_den[s]) if wtd_rate_den.get(s, 0.0) > 0 else 0.0
        result.append({
            "Row Labels": s,
            "_wired": d["wired"], "_paid_gal": d["paid_gal"],
            "_pulled": pulled, "_rem_alloc": rem_alloc,
            "_rem_inv": rem_inv, "_avg_cost": avg_cost,
            "_paid_back": sup_paid_back.get(s, 0.0),
            "_balance":   sup_balance.get(s, 0.0),
            "_wtd_rate":  wtd_rate,
        })

    total_wired     = sum(d["wired"]    for d in sup_data.values())
    total_paid_gal  = sum(d["paid_gal"] for d in sup_data.values())
    total_pulled    = sum(pulled_gal.values())
    total_rem_inv   = sum(rem_inv_gal.values())
    total_inv_mxn   = sum(rem_inv_mxn.values())
    total_avg_cost  = (total_inv_mxn / (total_rem_inv * 3.7854)) if total_rem_inv > 0 else 0.0
    total_rem_alloc = sum(rem_alloc_gal.values())
    total_paid_back = sum(sup_paid_back.values())
    total_balance   = sum(sup_balance.values())
    total_wtd_num   = sum(wtd_rate_num.values())
    total_wtd_den   = sum(wtd_rate_den.values())
    total_wtd_rate  = (total_wtd_num / total_wtd_den) if total_wtd_den > 0 else 0.0

    result.append({
        "Row Labels": "Total",
        "_wired": total_wired, "_paid_gal": total_paid_gal,
        "_pulled": total_pulled, "_rem_alloc": total_rem_alloc,
        "_rem_inv": total_rem_inv, "_avg_cost": total_avg_cost,
        "_paid_back": total_paid_back, "_balance": total_balance,
        "_wtd_rate": total_wtd_rate,
    })
    return result


# ── Inventory (from FIFO sheet + Purchase to BOL-RTB) ─────────────────────────
def _extract_inventory(wb):
    """
    Build hierarchy: bulk_plant -> product -> batch -> supplier -> invoice -> [bols]

    BOLs that span two batches/invoices (e.g. Batch="1 | 2", Invoice="A | B")
    are split into separate entries — one per batch/invoice — with proportional liters.
    Each BOL always appears INDIVIDUALLY under its own invoice.
    """
    # Step 1: read FIFO sheet for each RTB BOL
    ws_fifo = wb["FIFO"]
    fifo_headers = [str(v).strip() if v else f"col{j}"
                    for j, v in enumerate(next(ws_fifo.iter_rows(values_only=True)))]

    fifo_bols = {}  # bol_str -> {liters, remaining_l, cost_per_l, bp, prod}
    for row in ws_fifo.iter_rows(min_row=2, values_only=True):
        if not row[0]: break
        entry = {h: row[j] for j, h in enumerate(fifo_headers)}
        if entry.get("Type") != "RTB": continue
        bol = str(entry.get("BOL", "") or "")
        if not bol: continue
        liters = float(entry.get("Liters", 0) or 0)
        rem    = entry.get("Remaining L (BOL)")
        fifo_bols[bol] = {
            "liters":      liters,
            "remaining_l": float(rem) if rem is not None else liters,
            "cost_per_l":  float(entry.get("Cost / L (MXN)", 0) or 0),
            "bp":          str(entry.get("Bulk Plant", "") or ""),
            "prod":        str(entry.get("Product", "") or ""),
        }

    # Step 2: read Purchase to BOL-RTB for invoice, batch, supplier per BOL
    # For split BOLs (e.g. Batch="1 | 2", Invoice="A | B"), split into two entries
    bol_entries = []  # list of {bol, batch, invoice, supplier, alloc_frac}
    try:
        ws_bol = wb["Purchase to BOL-RTB"]
        for row in ws_bol.iter_rows(min_row=8, values_only=True):
            if not row[2]: break
            bol_str  = str(row[5]).strip() if row[5] else ""
            supplier = str(row[2]).strip() if row[2] else ""
            inv_raw  = str(row[3]).strip() if row[3] else ""
            bat_raw  = str(row[4]).strip() if row[4] else ""
            cost_j   = float(row[9])  if row[9]  is not None else 0.0  # J Cost/Gal USD
            cost_l_raw = str(row[11]).strip() if row[11] is not None else ""

            if not bol_str: continue

            # Split multi-batch/invoice BOLs into individual entries
            batches  = [b.strip() for b in bat_raw.split("|")] if "|" in bat_raw else [bat_raw]
            invoices = [v.strip() for v in inv_raw.split("|")] if "|" in inv_raw else [inv_raw]

            if len(batches) > 1 or len(invoices) > 1:
                # Cross-batch BOL: split proportionally
                # We don't have exact split fractions here, so split equally
                # (the FIFO sheet has the blended cost already)
                n = max(len(batches), len(invoices))
                for k in range(n):
                    b = batches[k] if k < len(batches) else batches[-1]
                    inv = invoices[k] if k < len(invoices) else invoices[-1]
                    bol_entries.append({
                        "bol": bol_str, "batch": b, "invoice": inv,
                        "supplier": supplier, "split_n": n, "split_k": k,
                    })
            else:
                bol_entries.append({
                    "bol": bol_str, "batch": bat_raw, "invoice": inv_raw,
                    "supplier": supplier, "split_n": 1, "split_k": 0,
                })
    except Exception as e:
        pass

    # Step 3: build hierarchy
    result = {}
    for entry in bol_entries:
        bol      = entry["bol"]
        batch    = entry["batch"]
        invoice  = entry["invoice"]
        supplier = entry["supplier"]
        split_n  = entry["split_n"]
        fifo     = fifo_bols.get(bol)
        if not fifo: continue

        bp   = fifo["bp"]
        prod = fifo["prod"]
        # Proportional liters for split BOLs
        liters      = round(fifo["liters"]      / split_n, 4)
        remaining_l = round(fifo["remaining_l"] / split_n, 4)
        cost_per_l  = fifo["cost_per_l"]  # blended cost same for all splits

        # Navigate: bp -> prod -> batch -> supplier -> invoice -> [bols]
        result.setdefault(bp, {})
        result[bp].setdefault(prod, {})
        result[bp][prod].setdefault(batch, {})
        result[bp][prod][batch].setdefault(supplier, {})
        result[bp][prod][batch][supplier].setdefault(invoice, [])
        result[bp][prod][batch][supplier][invoice].append({
            "bol":         bol,
            "liters":      liters,
            "remaining_l": remaining_l,
            "cost_per_l":  cost_per_l,
        })

    return result

# ── FIFO rows ──────────────────────────────────────────────────────────────────
def _extract_fifo(wb):
    ws = wb["FIFO"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v else f"col{j}" for j, v in enumerate(rows[0])]

    # Build BOL → supplier map from Purchase to BOL-RTB
    bol_to_sup = {}
    try:
        ws_bol = wb["Purchase to BOL-RTB"]
        bp_rows = list(ws_bol.iter_rows(values_only=True))
        bp_hdr_idx = next((i for i, r in enumerate(bp_rows) if r[0] and 'DashFuel' in str(r[0])), 6)
        bp_h = bp_rows[bp_hdr_idx]
        def _bc(name, default=None):
            for j, h in enumerate(bp_h):
                if h and name.lower() in str(h).lower(): return j
            return default
        col_b = _bc('BOL', 5)
        col_s = _bc('Supplier', 2)
        for row in bp_rows[bp_hdr_idx + 1:]:
            if not any(row): break
            bol_val = row[col_b] if col_b is not None else None
            sup_val = str(row[col_s] or '').strip() if col_s is not None else ''
            if bol_val and sup_val:
                bol_to_sup[str(bol_val)] = sup_val
    except Exception:
        pass

    result = []
    bol_idx = headers.index('BOL') if 'BOL' in headers else None
    for row in rows[1:]:
        if not row[0]: break
        entry = {}
        for j, h in enumerate(headers):
            v = row[j]
            if hasattr(v, 'isoformat'):
                v = v.strftime("%d/%m/%Y")
            elif isinstance(v, float):
                v = round(v, 4)
            entry[h] = v
        # Inject supplier from BOL lookup
        bol_val = str(row[bol_idx]) if bol_idx is not None and row[bol_idx] is not None else ''
        entry['Supplier'] = bol_to_sup.get(bol_val, '')
        result.append(entry)
    return result

# ── Meta ───────────────────────────────────────────────────────────────────────
def _extract_meta(wb, wb_fifo=None):
    os_rows = _extract_overall_summary(wb, wb_fifo=wb_fifo)
    total_os = next((r for r in os_rows if r.get("Row Labels","").upper().find("TOTAL") >= 0), {})
    return {
        "total_invoiced_usd":      total_os.get("_wired", 0),
        "total_gallons":           total_os.get("_paid_gal", 0),
        "total_wired":             total_os.get("_wired", 0),
        "paid_for_gallons":        total_os.get("_paid_gal", 0),
        "gallons_pulled":          total_os.get("_pulled", 0),
        "remaining_allocation":    total_os.get("_rem_alloc", 0),
        "remaining_inventory_gal": total_os.get("_rem_inv", 0),
        "avg_cost_inventory":      total_os.get("_avg_cost", 0),
        "amount_paid_back":        total_os.get("_paid_back", 0),
        "mexico_balance":          total_os.get("_balance", 0),
    }

# ── Investment Summary ─────────────────────────────────────────────────────────
def _extract_investment_summary(wb, wb_fifo=None, uploaded_at=None):
    """
    wb      = source workbook (original upload) — raw Load Tracking inputs
    wb_fifo = FIFO output workbook — has engine-written columns (Status, Remainder Amount MXN, FIFO sheet)
    """
    _wb_out = wb_fifo if wb_fifo is not None else wb  # for FIFO-written sheets
    f  = lambda v: float(v) if isinstance(v, (int, float)) else 0.0

    # ── 1. Committed Capital — user-entered, always cached ────────────────────
    ws_inv = wb["Investment Summary"]
    inv_rows = {i: list(row) for i, row in enumerate(ws_inv.iter_rows(values_only=True), start=1)}

    r3 = inv_rows.get(3, [])
    as_of_val = r3[6] if len(r3) > 6 else None
    if hasattr(as_of_val, "strftime"):
        as_of_str = as_of_val.strftime("%d-%b-%Y")
    elif as_of_val:
        as_of_str = str(as_of_val)
    else:
        as_of_str = uploaded_at or ""

    commits = []
    for ri in (7, 8):
        row = inv_rows.get(ri, [])
        if len(row) < 6: continue
        usd = f(row[2]); fx = f(row[5])
        dv  = row[4]
        dstr = dv.strftime("%d-%b-%y") if hasattr(dv, "strftime") else str(dv or "")
        mxn = usd * fx
        if usd > 0:
            commits.append({"round": str(row[1] or ""), "usd": usd, "mxn": mxn, "date": dstr, "fx": fx})

    total_committed_usd = sum(c["usd"] for c in commits)
    total_committed_mxn = sum(c["mxn"] for c in commits)
    avg_fx = total_committed_mxn / total_committed_usd if total_committed_usd else 17.31
    inv_share_pct = 0.40

    # ── 2. Allocation from Supplier Invoices — Status=ACTIVE rows ────────────
    # Col AA (27) = Status, Col Y (25) = Remainder Amount MXN, Col V (22) = Remainder Liters
    alloc_mxn = alloc_lit = 0.0
    try:
        ws_si_alloc = wb["Supplier Invoices"]  # source has cached Status/Remainder formulas
        si_alloc_rows = list(ws_si_alloc.iter_rows(values_only=True))
        si_hdr_idx = next((i for i, r in enumerate(si_alloc_rows)
                           if r[0] == "Batch" and len(r) > 2 and r[2] == "Supplier"), None)
        if si_hdr_idx is not None:
            sch = {str(v).strip(): j for j, v in enumerate(si_alloc_rows[si_hdr_idx]) if v}
            col_status  = sch.get("Status", 26)
            col_rem_mxn = sch.get("Remainder Amount MXN", 24)
            col_rem_lit = sch.get("Remainder Liters Paid and No BOL", 21)
            for row in si_alloc_rows[si_hdr_idx + 1:]:
                if not any(row): break
                status = str(row[col_status] or '').strip().upper()
                if status == 'ACTIVE':
                    alloc_mxn += f(row[col_rem_mxn])
                    alloc_lit += f(row[col_rem_lit])
    except Exception:
        pass

    def _parse_date(v):
        """Parse a single date value — datetime passthrough or string parse."""
        if hasattr(v, 'date'): return v
        if isinstance(v, str):
            from datetime import datetime as _dt
            s = v.strip()
            for fmt in ('%m/%d/%Y','%d/%m/%Y','%Y-%m-%d'):
                try: return _dt.strptime(s, fmt)
                except: pass
        return None

    def _resolve_dates_contextual(raw_dates):
        """
        Given a list of raw date values (datetime | str | None) in load order,
        resolve ambiguous MM/DD vs DD/MM strings using context from neighbors.
        Rule: dates should be monotonically increasing (loads are in order).
        If a string has day > 12 → must be DD/MM, flip to MM/DD interpretation.
        If both interpretations valid → pick the one closer to neighbors.
        """
        from datetime import datetime as _dt

        def _candidates(v):
            """Return (mm_dd, dd_mm) candidate datetimes for a string, or (dt, dt) if unambiguous."""
            if hasattr(v, 'date'): return (v, v)
            if not isinstance(v, str): return (None, None)
            s = v.strip()
            parts = s.replace('-','/').split('/')
            if len(parts) != 3: return (None, None)
            try:
                a, b, yr = int(parts[0]), int(parts[1]), int(parts[2])
            except: return (None, None)
            # Try MM/DD/YYYY and DD/MM/YYYY
            mm_dd = dd_mm = None
            try: mm_dd = _dt(yr, a, b)  # MM/DD
            except: pass
            try: dd_mm = _dt(yr, b, a)  # DD/MM
            except: pass
            # If day > 12, only DD/MM is valid
            if a > 12: return (dd_mm, dd_mm)
            if b > 12: return (mm_dd, mm_dd)
            return (mm_dd, dd_mm)

        # First pass: get unambiguous anchors
        resolved = [None] * len(raw_dates)
        candidates = [_candidates(v) for v in raw_dates]

        # Mark unambiguous ones
        for i, (mm, dd) in enumerate(candidates):
            if mm == dd and mm is not None:
                resolved[i] = mm

        # Second pass: resolve ambiguous ones using nearest resolved neighbor
        for i, (mm, dd) in enumerate(candidates):
            if resolved[i] is not None: continue
            if mm is None and dd is None: continue
            if mm is None: resolved[i] = dd; continue
            if dd is None: resolved[i] = mm; continue
            # Both valid — find nearest resolved neighbor and pick closer interpretation
            neighbor = None
            for j in range(1, len(raw_dates)):
                if i-j >= 0 and resolved[i-j] is not None:
                    neighbor = resolved[i-j]; break
                if i+j < len(raw_dates) and resolved[i+j] is not None:
                    neighbor = resolved[i+j]; break
            if neighbor is None:
                # No context — default to MM/DD (US format)
                resolved[i] = mm
            else:
                # Pick whichever is closer to neighbor and not wildly out of range
                diff_mm = abs((mm - neighbor).days)
                diff_dd = abs((dd - neighbor).days)
                resolved[i] = mm if diff_mm <= diff_dd else dd

        return resolved

    # ── 3. Load Tracking — replicate Table6 formulas ──────────────────────────
    ws_lt = wb["Load Tracking"]
    lt_rows = list(ws_lt.iter_rows(values_only=True))
    lt_h = {str(v).strip(): i for i, v in enumerate(lt_rows[0]) if v}

    def _get(row, col_name, default=0.0):
        idx = lt_h.get(col_name)
        if idx is None: return default
        return f(row[idx])

    rtb_total_lit = 0.0
    btc_total_lit = 0.0
    btc_pend_mxn = btc_pend_lit = 0.0
    rtc_pend_mxn = rtc_pend_lit = 0.0
    rec_btc_tc = rec_btc_sale = rec_btc_margin = rec_btc_lit = 0.0
    rec_rtc_tc = rec_rtc_sale = rec_rtc_margin = rec_rtc_lit = 0.0

    # Build BOL → batch + liters from RTB rows for proportional batch splitting
    bol_batch_map = {}  # bol → {batch, liters}
    for row in lt_rows[1:]:
        if not row[0]: continue
        if str(row[lt_h.get('Customer Groups',10)] or '').strip() != 'RTB': continue
        bol   = str(row[lt_h.get('BOL Number',30)] or '').strip()
        batch = str(row[lt_h.get('Batch',57)] or '').strip()
        liters_rtb = _get(row, 'Delivered Net Liters') or _get(row, 'Net Liters')
        if bol and liters_rtb > 0:
            bol_batch_map[bol] = {'batch': batch, 'liters': liters_rtb}

    # Build RTB dates map using contextually resolved dates
    rtb_date_rows = [(row, lt_h.get('BOL Number',30), lt_h.get('Pickup Date',3))
                     for row in lt_rows[1:]
                     if row[0] and str(row[lt_h.get('Customer Groups',10)] or '').strip() == 'RTB']
    rtb_raw_dates   = [row[lt_h.get('Pickup Date',3)] for row, _, _ in rtb_date_rows]
    rtb_resolved    = _resolve_dates_contextual(rtb_raw_dates)
    rtb_dates_map   = {}
    for (row, col_bol, _), resolved_dt in zip(rtb_date_rows, rtb_resolved):
        bol = str(row[col_bol] or '').strip()
        if bol and resolved_dt:
            rtb_dates_map[bol] = resolved_dt

    # Resolve BTC pickup dates contextually too
    btc_date_rows = [row for row in lt_rows[1:]
                     if row[0] and str(row[lt_h.get('Customer Groups',10)] or '').strip() == 'BTC']
    btc_raw_pickups   = [row[lt_h.get('Pickup Date',3)] for row in btc_date_rows]
    btc_resolved_pkup = _resolve_dates_contextual(btc_raw_pickups)
    # keyed by sequential BTC index (0,1,2...)
    btc_pickup_by_idx = {i: dt for i, dt in enumerate(btc_resolved_pkup)}

    def _compute_batch_splits(bol_source_str, bol_batch_map):
        """Returns {batch_num: fraction} for a BTC load based on liters from each batch."""
        if not bol_source_str: return {}
        src_bols = [b.strip() for b in bol_source_str.split('|')]
        batch_liters = {}
        for bol in src_bols:
            if bol not in bol_batch_map: continue
            info = bol_batch_map[bol]
            # Normalize batch — take first part if "1 | 2" style
            b = info['batch'].split('|')[0].strip()
            batch_liters[b] = batch_liters.get(b, 0.0) + info['liters']
        total = sum(batch_liters.values())
        if not total: return {}
        return {b: round(l/total, 6) for b, l in batch_liters.items()}

    btc_loads = []
    cycle_days_list = []
    btc_row_counter = 0  # index into btc_pickup_map

    col_payment_date = lt_h.get('Payment Date', 61)
    col_bol_source   = lt_h.get('BOL Source', 59)

    for row in lt_rows[1:]:
        if not row[0]: continue
        typ    = str(row[lt_h.get('Customer Groups', 10)] or '').strip()
        status = str(row[lt_h.get('Invoice Status', 55)] or '').strip().upper()
        liters = _get(row, 'Delivered Net Liters')
        if liters <= 0:
            liters = _get(row, 'Net Liters')

        supply_cost_l = _get(row, 'Supply Cost')
        freight       = _get(row, 'Freight Cost')
        commission    = _get(row, 'COMISSION')
        extra         = _get(row, 'EXTRA')
        fuel_cost     = _get(row, 'Fuel Cost')

        freight_l    = freight / liters    if liters > 0 else 0.0
        comm_l       = commission / liters if liters > 0 else 0.0
        extra_l      = extra / liters      if liters > 0 else 0.0
        total_cost_l = supply_cost_l + freight_l + comm_l + extra_l
        total_cost   = total_cost_l * liters

        if typ in ('BTC', 'RTC'):
            price_l    = (fuel_cost / liters) + freight_l + comm_l + extra_l if liters > 0 else 0.0
        else:
            price_l    = 0.0

        total_sale   = price_l * liters
        total_margin = total_sale - total_cost

        if typ == 'RTB':
            rtb_total_lit  += liters
        elif typ == 'BTC':
            btc_total_lit  += liters
            pickup         = btc_pickup_by_idx.get(btc_row_counter)
            btc_row_counter += 1
            payment_date   = _parse_date(row[col_payment_date]) if col_payment_date is not None else None
            bol_source_str = str(row[col_bol_source] or '') if col_bol_source is not None else ''
            load_id        = str(row[0] or '')
            customer       = str(row[lt_h.get('Customer',8)] or '')
            pickup_str     = pickup.strftime('%m/%d/%Y') if pickup else ''
            payment_str    = payment_date.strftime('%m/%d/%Y') if payment_date else ''

            # Cash conversion cycle: earliest source RTB date → payment date
            cycle = None
            if payment_date and bol_source_str:
                src_bols = [b.strip() for b in bol_source_str.split('|')]
                src_dates = [rtb_dates_map[b] for b in src_bols if b in rtb_dates_map]
                if src_dates:
                    cycle = (payment_date - min(src_dates)).days
                    if 0 <= cycle <= 90:   # sanity filter
                        cycle_days_list.append(cycle)

            btc_loads.append({
                'load':        load_id,
                'customer':    customer,
                'date':        pickup_str,
                'payment_date': payment_str,
                'liters':      liters,
                'price_l':     round(total_sale / liters, 4) if liters else 0,
                'cost_l':      round(total_cost_l, 4),
                'margin_l':    round((total_sale - total_cost) / liters, 4) if liters else 0,
                'margin_pct':  _get(row, 'Margin %'),
                'total_margin': round(total_sale - total_cost, 2),
                'total_sale':  round(total_sale, 2),
                'total_cost':  round(total_cost, 2),
                'status':      str(row[lt_h.get('Invoice Status',56)] or ''),
                'cycle_days':  cycle,
                'bol_source':  bol_source_str,
                'batch':       str(row[lt_h.get('Batch Source', 60)] or '').strip(),
                'batch_splits': _compute_batch_splits(bol_source_str, bol_batch_map),
            })

            if status == 'PAID':
                rec_btc_tc     += total_cost
                rec_btc_sale   += total_sale
                rec_btc_margin += total_sale - total_cost
                rec_btc_lit    += liters
            else:
                btc_pend_mxn   += total_sale
                btc_pend_lit   += liters
        elif typ == 'RTC':
            if status == 'PAID':
                rec_rtc_tc     += total_cost
                rec_rtc_sale   += total_sale
                rec_rtc_margin += total_margin
                rec_rtc_lit    += liters
            else:
                rtc_pend_mxn   += total_sale
                rtc_pend_lit   += liters

    avg_cycle_days = round(sum(cycle_days_list) / len(cycle_days_list), 1) if cycle_days_list else None

    # Use FIFO sheet remaining_l × cost_per_l for accurate FIFO-weighted cost
    inv_lit = inv_mxn = 0.0
    try:
        ws_fifo_inv = _wb_out["FIFO"]
        fifo_inv_rows = list(ws_fifo_inv.iter_rows(values_only=True))
        fh = {str(v).strip(): j for j, v in enumerate(fifo_inv_rows[0]) if v}
        for row in fifo_inv_rows[1:]:
            if not row[0]: break
            if str(row[fh.get('Type', 2)]).strip() != 'RTB': continue
            rem_l  = f(row[fh.get('Remaining L (BOL)', 13)])
            cost_l = f(row[fh.get('Cost / L (MXN)', 9)])
            inv_lit += rem_l
            inv_mxn += rem_l * cost_l
    except Exception:
        # Fallback: simple RTB - BTC with avg cost if FIFO sheet unavailable
        inv_lit = max(0.0, rtb_total_lit - btc_total_lit)

    # ── 4. KPIs ───────────────────────────────────────────────────────────────
    active_capital = alloc_mxn + inv_mxn + btc_pend_mxn + rtc_pend_mxn
    available      = total_committed_mxn - active_capital
    recovered_mxn  = rec_btc_sale + rec_rtc_sale
    revolved       = recovered_mxn / total_committed_mxn if total_committed_mxn else 0.0
    total_margin   = rec_btc_margin + rec_rtc_margin
    investor_share = total_margin * inv_share_pct

    # ── Stage days computation ─────────────────────────────────────────────────
    # Uses Purchase to BOL-RTB for RTB dates, SI for wire dates, LT for BTC/payment dates
    # All stages: use today for open-ended (still in allocation/inventory/pending)
    from datetime import datetime as _dt_now
    _TODAY = _dt_now.now()

    # Wire dates per invoice from SI
    wire_date_by_inv = {}
    try:
        ws_si_wd = wb["Supplier Invoices"]
        si_wd_rows = list(ws_si_wd.iter_rows(values_only=True))
        si_wd_hdr  = next((i for i,r in enumerate(si_wd_rows) if r[0]=='Batch' and r[2]=='Supplier'), None)
        if si_wd_hdr is not None:
            sc_wd = {str(v).strip():j for j,v in enumerate(si_wd_rows[si_wd_hdr]) if v}
            for row in si_wd_rows[si_wd_hdr+1:]:
                if not any(row): break
                inv_k = str(row[sc_wd.get('Invoice #',3)] or '').strip()
                dt_w  = _parse_date(row[sc_wd.get('Date',1)])
                if inv_k and dt_w:
                    wire_date_by_inv[inv_k] = dt_w
    except Exception: pass

    # RTB dates and gallons per BOL from Purchase to BOL-RTB
    bol_rtb_info = {}  # bol → {rtb_date, gallons, invoice}
    try:
        ws_bol_sd = wb["Purchase to BOL-RTB"] if wb_fifo is None else (wb_fifo or wb)["Purchase to BOL-RTB"]
        # Use source wb for raw dates
        ws_bol_sd = wb["Purchase to BOL-RTB"]
        bp_sd_rows = list(ws_bol_sd.iter_rows(values_only=True))
        bp_sd_hdr  = next((i for i,r in enumerate(bp_sd_rows) if r[0] and 'DashFuel' in str(r[0])), 6)
        bh_sd = {str(v).strip():j for j,v in enumerate(bp_sd_rows[bp_sd_hdr]) if v}
        for row in bp_sd_rows[bp_sd_hdr+1:]:
            if not any(row): break
            bol_k = str(row[bh_sd.get('BOL',5)] or '').strip()
            inv_k = str(row[bh_sd.get('Supplier Invoice',3)] or '').strip()
            gal_k = f(row[bh_sd.get('Gallons',6)])
            dt_k  = _parse_date(row[bh_sd.get('Date',1)])
            if bol_k:
                bol_rtb_info[bol_k] = {'rtb_date': dt_k, 'gallons': gal_k, 'invoice': inv_k}
    except Exception: pass

    # ── 1. Allocation days: wire_date → rtb_date (or today if still in alloc) ──
    alloc_num = alloc_den = 0.0
    for bol, info in bol_rtb_info.items():
        rtb_dt = info['rtb_date']
        gal    = info['gallons']
        inv_s  = info['invoice']
        end_dt = rtb_dt if rtb_dt else _TODAY
        # Find wire date — use first matching part of split invoice
        wire_dt = None
        for part in [p.strip() for p in inv_s.split('|')]:
            if part in wire_date_by_inv:
                wire_dt = wire_date_by_inv[part]
                break
        if wire_dt and gal > 0:
            days = max(0, (end_dt - wire_dt).days)
            alloc_num += gal * days
            alloc_den += gal

    # ── 2. Inventory days: rtb_date → btc_pickup_date (or today if still in inv) ──
    inv_sd_num = inv_sd_den = 0.0
    # From BTCs: each source BOL contributed gallons, rtb→btc_pickup
    for load in btc_loads:
        pickup_dt = _parse_date(load['date']) if load['date'] else None
        if not pickup_dt: continue
        gals_load = load['liters'] / 3.785411784 if load['liters'] else 0
        bol_src_str = load.get('bol_source', '')
        src_bols = [b.strip() for b in bol_src_str.split('|') if b.strip()] if bol_src_str else []
        if src_bols:
            gal_per = gals_load / len(src_bols)
            for b in src_bols:
                if b in bol_rtb_info and bol_rtb_info[b]['rtb_date']:
                    days = max(0, (pickup_dt - bol_rtb_info[b]['rtb_date']).days)
                    inv_sd_num += gal_per * days
                    inv_sd_den += gal_per
    # Still in inventory: use today
    try:
        ws_fifo_sd = _wb_out["FIFO"]
        fifo_sd_rows = list(ws_fifo_sd.iter_rows(values_only=True))
        fh_sd = {str(v).strip():j for j,v in enumerate(fifo_sd_rows[0]) if v}
        for row in fifo_sd_rows[1:]:
            if not row[0]: break
            if str(row[fh_sd.get('Type',2)]).strip() != 'RTB': continue
            rem_l = f(row[fh_sd.get('Remaining L (BOL)',13)])
            if rem_l <= 0: continue
            bol_k = str(row[fh_sd.get('BOL',7)] or '').strip()
            rem_gal = rem_l / 3.785411784
            if bol_k in bol_rtb_info and bol_rtb_info[bol_k]['rtb_date']:
                days = max(0, (_TODAY - bol_rtb_info[bol_k]['rtb_date']).days)
                inv_sd_num += rem_gal * days
                inv_sd_den += rem_gal
    except Exception: pass

    # ── 3. Pending collection days: btc_pickup → payment (or today) ──
    pend_sd_num = pend_sd_den = 0.0
    for load in btc_loads:
        pickup_dt  = _parse_date(load['date']) if load['date'] else None
        payment_dt = _parse_date(load['payment_date']) if load['payment_date'] else None
        gals_load  = load['liters'] / 3.785411784 if load['liters'] else 0
        if pickup_dt and gals_load > 0:
            end_dt = payment_dt if payment_dt else _TODAY
            days = max(0, (end_dt - pickup_dt).days)
            pend_sd_num += gals_load * days
            pend_sd_den += gals_load

    avg_alloc_days = round(alloc_num / alloc_den, 1) if alloc_den > 0 else None
    avg_inv_days   = round(inv_sd_num / inv_sd_den, 1) if inv_sd_den > 0 else None
    avg_pend_days  = round(pend_sd_num / pend_sd_den, 1) if pend_sd_den > 0 else None
    avg_cycle_days = round(sum(cycle_days_list) / len(cycle_days_list), 1) if cycle_days_list else None
    paid_loads     = [l for l in btc_loads if l['status'].upper() == 'PAID']
    avg_margin_per_load = (sum(l['total_margin'] for l in paid_loads) / len(paid_loads)) if paid_loads else 0
    avg_liters_per_load = (sum(l['liters'] for l in paid_loads) / len(paid_loads)) if paid_loads else 0
    loads_remaining     = round(inv_lit / avg_liters_per_load, 1) if avg_liters_per_load else 0
    btc_pickup_dates    = sorted([_parse_date(l['date']) for l in btc_loads if l['date']])
    avg_days_between_loads = 0
    if len(btc_pickup_dates) > 1:
        gaps = [(btc_pickup_dates[i]-btc_pickup_dates[i-1]).days for i in range(1,len(btc_pickup_dates))]
        avg_days_between_loads = round(sum(gaps)/len(gaps), 1)
    proj_profit          = round(loads_remaining * avg_margin_per_load, 2)
    proj_total_profit    = round(total_margin + proj_profit, 2)
    proj_investor_return = round(proj_total_profit * inv_share_pct, 2)
    paid_margins = [l['total_margin'] for l in paid_loads]
    first3_avg   = round(sum(paid_margins[:3])/3, 2) if len(paid_margins) >= 3 else None
    last3_avg    = round(sum(paid_margins[-3:])/3, 2) if len(paid_margins) >= 3 else None

    def _rec(tc, sale, margin, liters):
        roi   = margin / tc     if tc     > 0 else 0.0
        mxnl  = margin / liters if liters > 0 else 0.0
        usdg  = mxnl * 3.7854 / avg_fx if avg_fx > 0 else 0.0
        return {"tc": tc, "sale": sale, "liters": liters, "margin": margin,
                "roi": roi, "mxnl": mxnl, "usdgal": usdg}

    btc_rec = _rec(rec_btc_tc, rec_btc_sale, rec_btc_margin, rec_btc_lit)
    rtc_rec = _rec(rec_rtc_tc, rec_rtc_sale, rec_rtc_margin, rec_rtc_lit)
    tot_lit = rec_btc_lit + rec_rtc_lit
    tot_tc  = rec_btc_tc  + rec_rtc_tc
    tot_rec = _rec(tot_tc, recovered_mxn, total_margin, tot_lit)

    return {
        "as_of":               as_of_str,
        "commits":             commits,
        "total_committed_usd": total_committed_usd,
        "total_committed_mxn": total_committed_mxn,
        "active_capital":      active_capital,
        "available":           available,
        "recovered_mxn":       recovered_mxn,
        "revolved":            revolved,
        "total_margin":        total_margin,
        "investor_share":      investor_share,
        "inv_share_pct":       inv_share_pct,
        "active_detail": {
            "alloc_mxn":       alloc_mxn,
            "alloc_liters":    alloc_lit,
            "inv_mxn":         inv_mxn,
            "inv_liters":      inv_lit,
            "rtc_pend_mxn":    rtc_pend_mxn,
            "rtc_pend_liters": rtc_pend_lit,
            "btc_pend_mxn":    btc_pend_mxn,
            "btc_pend_liters": btc_pend_lit,
            "total_mxn":       active_capital,
            "total_liters":    alloc_lit + inv_lit + btc_pend_lit + rtc_pend_lit,
        },
        "recovered_detail": {
            "btc":   btc_rec,
            "rtc":   rtc_rec,
            "total": tot_rec,
        },
        "avg_cycle_days":          avg_cycle_days,
        "avg_days_between_loads":  avg_days_between_loads,
        "avg_alloc_days":          avg_alloc_days,
        "avg_inv_days":            avg_inv_days,
        "avg_pend_days":           avg_pend_days,
        "btc_loads":               btc_loads,
        "projections": {
            "loads_remaining":      loads_remaining,
            "avg_margin_per_load":  round(avg_margin_per_load, 2),
            "proj_profit":          proj_profit,
            "proj_total_profit":    proj_total_profit,
            "proj_investor_return": proj_investor_return,
            "first3_avg_margin":    first3_avg,
            "last3_avg_margin":     last3_avg,
            "inv_liters":           round(inv_lit, 0),
        },
    }


# ── Purchase BOLs tab ──────────────────────────────────────────────────────────
def _extract_bol(wb, wb_src=None):
    try:
        ws = wb["Purchase to BOL-RTB"]
        rows = list(ws.iter_rows(values_only=True))
        # Also load source rows for fallback (freight, invoice amounts)
        src_rows = None
        if wb_src is not None:
            try:
                ws_src = wb_src["Purchase to BOL-RTB"]
                src_rows = list(ws_src.iter_rows(values_only=True))
            except Exception:
                pass

        # Find header row
        header_row = None
        for i, row in enumerate(rows):
            if row[0] and 'DashFuel' in str(row[0]):
                header_row = i
                break
        if header_row is None:
            for i, row in enumerate(rows):
                if any('DashFuel' in str(v) for v in row if v):
                    header_row = i
                    break
        if header_row is None:
            return {"rows": []}

        headers = rows[header_row]
        def _col(name, default=None):
            for j, h in enumerate(headers):
                if h and name.lower() in str(h).lower():
                    return j
            return default

        col_bol          = _col('BOL', 5)
        col_dashfuel     = _col('DashFuel Number', 0)
        col_supplier     = _col('Supplier', 2)
        col_inv          = _col('Supplier Invoice', 3)
        col_batch        = _col('Batch', 4)
        col_date         = _col('Date', 1)
        col_gal          = _col('Gallons', 6)
        col_lit          = _col('Liters', 7)
        col_prod         = _col('Product', 8)
        col_cost_gal     = _col('Total Cost /Gal', 16)
        col_adder        = _col('Adder', 10)
        col_total_gal    = _col('Cost/Gal + Adder', 11)
        col_mxnl         = _col('Supply Cost DashFuel', 12)
        col_carrier      = _col('Carrier', 13)
        col_freight      = _col('Freight/Load', 14)
        col_freight_gal  = _col('Freight/Gal', 15)
        col_inv_num      = _col('Invoice #', 17)
        col_inv_amt      = _col('Invoice Amount', 18)
        col_customer     = _col('MX Customer', 19)
        col_received     = _col('Received Payments', 20)
        col_balance      = _col('Balance', 21)
        col_fx_rate      = _col('FX Payment', 25)

        fv = lambda v: round(float(v), 4) if isinstance(v, (int, float)) else None
        f  = lambda v: float(v) if isinstance(v, (int, float)) else 0.0

        # Load source rows for formula columns that are None in the FIFO output:
        # col0 (DashFuel#), col16 (Total Cost/Gal), col18 (Invoice Amount), col21 (Balance)
        src_rows = None
        if wb_src is not None:
            try:
                src_rows = list(wb_src["Purchase to BOL-RTB"].iter_rows(values_only=True))
            except Exception:
                pass

        def _src(row_idx, col):
            """Read a column from the source row as fallback for formula columns."""
            if src_rows is None or col is None:
                return None
            src_idx = header_row + 1 + row_idx
            if src_idx >= len(src_rows):
                return None
            return src_rows[src_idx][col]

        result_rows = []
        for row_idx, row in enumerate(rows[header_row + 1:]):
            if not row[0] and not (col_supplier and row[col_supplier]):
                if not any(v for v in row):
                    break
                continue

            gallons = f(row[col_gal]) if col_gal is not None else 0.0
            liters  = gallons * 3.785411784

            # For formula columns, fall back to source if FIFO output has None
            inv_num     = str(row[col_inv_num] or '') if col_inv_num is not None else ''
            invoice_amt = fv(row[col_inv_amt]  if (col_inv_amt  is not None and row[col_inv_amt]  is not None) else _src(row_idx, col_inv_amt))
            received    = fv(row[col_received]) if col_received is not None else None
            balance     = fv(row[col_balance]   if (col_balance  is not None and row[col_balance]  is not None) else _src(row_idx, col_balance))
            cost_gal    = fv(row[col_cost_gal]  if (col_cost_gal is not None and row[col_cost_gal] is not None) else _src(row_idx, col_cost_gal))
            dashfuel    = str(row[col_dashfuel] or _src(row_idx, col_dashfuel) or '') if col_dashfuel is not None else ''

            # Date in English format mm/dd/YYYY
            raw_date = row[col_date] if col_date is not None else None
            if hasattr(raw_date, 'strftime'):
                date_str = raw_date.strftime("%m/%d/%Y")
            elif raw_date:
                date_str = str(raw_date)
            else:
                date_str = ''

            r = {
                "date":         date_str,
                "dashfuel_num": dashfuel,
                "product":      str(row[col_prod] or '')     if col_prod     is not None else '',
                "bol":          str(row[col_bol] or '')      if col_bol      is not None else '',
                "gallons":      gallons    if gallons else None,
                "liters":       liters     if gallons else None,
                "cost_gal_usd": cost_gal,
                "fx_rate":      fv(row[col_fx_rate])         if col_fx_rate  is not None else None,
                "invoice_num":  inv_num,
                "invoice_amt":  invoice_amt,
                "received":     received,
                "balance":      balance,
                "supplier":     str(row[col_supplier] or '') if col_supplier is not None else '',
                "inv_num":      str(row[col_inv] or '')      if col_inv      is not None else '',
                "batch":        str(row[col_batch] or '')    if col_batch    is not None else '',
                "adder":        fv(row[col_adder])           if col_adder    is not None else None,
                "total_gal":    fv(row[col_total_gal])       if col_total_gal is not None else None,
                "mxn_l":        fv(row[col_mxnl])            if col_mxnl     is not None else None,
                "carrier":      str(row[col_carrier] or '')  if col_carrier  is not None else '',
                "freight":      fv(row[col_freight])         if col_freight  is not None else None,
                "customer":     str(row[col_customer] or '') if col_customer is not None else '',
            }
            if r["bol"] or r["supplier"]:
                result_rows.append(r)

        # Compute summary KPIs
        f2 = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
        total_invoiced    = sum(f2(r["invoice_amt"]) for r in result_rows if r.get("invoice_amt"))
        received_payments = sum(f2(r["received"])    for r in result_rows if r.get("received"))
        # Open balance only for rows that have an invoice number
        open_balance      = sum(f2(r["balance"]) for r in result_rows if r.get("invoice_num") and f2(r.get("balance", 0)) > 0)

        # Not invoiced yet: BOL rows with no invoice number — compute directly
        total_not_invoiced = sum(
            f2(r["invoice_amt"]) for r in result_rows
            if not r.get("invoice_num") and r.get("invoice_amt")
        )

        # Add status to each row
        for r in result_rows:
            if r.get("invoice_amt") and r.get("invoice_num"):
                r["status"] = "paid" if (r.get("balance") or 0) <= 0.01 else "open"
            else:
                r["status"] = "not_invoiced"

        return {
            "rows": result_rows,
            "summary": {
                "total_invoiced":    total_invoiced,
                "received_payments": received_payments,
                "open_balance":      open_balance,
                "total_not_invoiced": total_not_invoiced,
            }
        }
    except Exception as e:
        return {"rows": [], "summary": {}}


# ── Overview / How Capital Works ───────────────────────────────────────────────
def _extract_overview(wb):
    try:
        ws = wb["Overview"]
        rows = list(ws.iter_rows(values_only=True))
        steps = []
        glossary = []
        in_glossary = False
        for i, row in enumerate(rows, start=1):
            if not any(v for v in row): continue
            label = str(row[1]).strip() if row[1] else ""
            text  = str(row[2]).strip() if row[2] else ""
            if label == "GLOSSARY":
                in_glossary = True
                continue
            if label == "HOW YOUR CAPITAL WORKS": continue
            if "This document" in label: continue
            if not label: continue
            if in_glossary and text:
                glossary.append({"term": label, "definition": text})
            elif not in_glossary and text:
                steps.append({"title": label, "body": text})
        return {"steps": steps, "glossary": glossary}
    except:
        return {"steps": [], "glossary": []}
